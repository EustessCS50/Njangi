import csv
import io
from datetime import datetime
from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
except Exception:
    canvas = None


def queryset_to_csv_response(queryset, field_names, filename=None):
    filename = filename or f"export-{datetime.utcnow().date()}.csv"
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(field_names)
    for obj in queryset:
        row = []
        for f in field_names:
            val = getattr(obj, f, '')
            # call if callable
            try:
                if callable(val):
                    val = val()
            except Exception:
                val = ''
            row.append(val)
        writer.writerow(row)
    resp = HttpResponse(output.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def queryset_to_xlsx_response(queryset, field_names, filename=None):
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed')
    filename = filename or f"export-{datetime.utcnow().date()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(field_names)
    for obj in queryset:
        row = []
        for f in field_names:
            val = getattr(obj, f, '')
            try:
                if callable(val):
                    val = val()
            except Exception:
                val = ''
            row.append(val)
        ws.append(row)
    # auto width
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                l = len(str(cell.value))
                if l > max_length:
                    max_length = l
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max(10, max_length + 2)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def group_report_xlsx(members_qs, filename=None):
    """Generate a nicer XLSX workbook with two sheets: Members and Summary."""
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed')
    filename = filename or f"group-report-{datetime.utcnow().date()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'
    headers = ['First name', 'Last name', 'Phone', 'Email', 'Joined', 'Total savings', 'Outstanding loans']
    ws.append(headers)
    from . import exporters as _dummy  # keep style
    for m in members_qs:
        try:
            total_savings = m.total_savings() if hasattr(m, 'total_savings') else ''
        except Exception:
            total_savings = ''
        try:
            outstanding = m.outstanding_loans() if hasattr(m, 'outstanding_loans') else ''
        except Exception:
            outstanding = ''
        ws.append([m.first_name, m.last_name, m.phone, m.email, str(m.joined_at), str(total_savings), str(outstanding)])
    # summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Metric', 'Value'])
    total_members = members_qs.count()
    total_contrib = 0
    try:
        from django.db.models import Sum as _Sum
        total_contrib = 0
    except Exception:
        total_contrib = 0
    ws2.append(['Total members', total_members])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def member_statement_pdf_response(member, meeting=None, filename=None):
    if canvas is None:
        raise RuntimeError('reportlab not installed')
    filename = filename or f"statement-{member.first_name}-{member.last_name or ''}.pdf"
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    left = 40
    top = height - 40
    line_height = 14

    def new_page():
        nonlocal c, left, top, y
        if 'c' in locals():
            c.showPage()
        y = height - 40

    # Header
    c.setFont('Helvetica-Bold', 16)
    c.drawString(left, top, f"Member Statement: {member}")
    c.setFont('Helvetica', 10)
    y = top - 28

    # Summary block
    stmt = member.statement(meeting) if hasattr(member, 'statement') else {
        'total_savings': 0,
        'outstanding_loans': 0,
        'eligibility_suggestion': 0,
    }
    c.drawString(left, y, f"Total savings: {stmt['total_savings']}")
    y -= line_height
    c.drawString(left, y, f"Outstanding loans: {stmt['outstanding_loans']}")
    y -= line_height
    c.drawString(left, y, f"Eligibility suggestion: {stmt['eligibility_suggestion']}")
    y -= line_height * 2

    # Loans section
    c.setFont('Helvetica-Bold', 12)
    c.drawString(left, y, 'Loans')
    y -= line_height
    c.setFont('Helvetica-Bold', 10)
    c.drawString(left, y, 'ID')
    c.drawString(left + 50, y, 'Principal')
    c.drawString(left + 140, y, 'Rate')
    c.drawString(left + 200, y, 'Status')
    c.drawString(left + 300, y, 'Due')
    c.drawString(left + 380, y, 'Outstanding')
    y -= line_height
    c.setFont('Helvetica', 10)

    loans = []
    try:
        for ms in member.memberships.select_related('meeting').all():
            for loan in ms.loans.all().order_by('-created_at'):
                try:
                    out = loan.outstanding()
                except Exception:
                    out = 'n/a'
                loans.append((loan.id, loan.principal, loan.interest_rate, loan.status, loan.due_date, out))
    except Exception:
        loans = []

    if not loans:
        c.drawString(left, y, 'No loans')
        y -= line_height
    else:
        for lid, principal, rate, status, due, outstanding_amt in loans:
            if y < 80:
                new_page()
            c.drawString(left, y, str(lid))
            c.drawString(left + 50, y, str(principal))
            c.drawString(left + 140, y, str(rate))
            c.drawString(left + 200, y, str(status))
            c.drawString(left + 300, y, str(due or ''))
            c.drawString(left + 380, y, str(outstanding_amt))
            y -= line_height

    y -= line_height

    # Repayments section
    c.setFont('Helvetica-Bold', 12)
    c.drawString(left, y, 'Recent Repayments')
    y -= line_height
    c.setFont('Helvetica-Bold', 10)
    c.drawString(left, y, 'Date')
    c.drawString(left + 90, y, 'Amount')
    c.drawString(left + 170, y, 'Loan ID')
    c.drawString(left + 230, y, 'Note')
    y -= line_height
    c.setFont('Helvetica', 10)

    repayments = []
    try:
        for ms in member.memberships.all():
            for loan in ms.loans.all():
                for rep in loan.repayments.order_by('-date')[:20]:
                    repayments.append((rep.date, rep.amount, getattr(loan, 'id', ''), rep.note))
    except Exception:
        repayments = []

    if not repayments:
        c.drawString(left, y, 'No repayments')
        y -= line_height
    else:
        for date, amount, loan_id, note in repayments:
            if y < 80:
                new_page()
            c.drawString(left, y, str(date))
            c.drawString(left + 90, y, str(amount))
            c.drawString(left + 170, y, str(loan_id))
            c.drawString(left + 230, y, str(note)[:40])
            y -= line_height

    y -= line_height

    # Transactions combined table (contributions + repayments) sorted by date desc
    c.setFont('Helvetica-Bold', 12)
    c.drawString(left, y, 'Transactions (Recent)')
    y -= line_height
    c.setFont('Helvetica-Bold', 10)
    c.drawString(left, y, 'Date')
    c.drawString(left + 90, y, 'Type')
    c.drawString(left + 170, y, 'Amount')
    c.drawString(left + 260, y, 'Related')
    y -= line_height
    c.setFont('Helvetica', 10)

    txs = []
    try:
        for ms in member.memberships.select_related('meeting').all():
            for cobj in ms.contributions.order_by('-date')[:50]:
                txs.append((cobj.date, 'contribution', cobj.amount, str(ms.meeting)))
            for loan in ms.loans.all():
                for rep in loan.repayments.order_by('-date')[:50]:
                    txs.append((rep.date, 'repayment', rep.amount, f'Loan {loan.id}'))
    except Exception:
        txs = []

    # sort by date desc
    try:
        txs_sorted = sorted(txs, key=lambda t: t[0], reverse=True)
    except Exception:
        txs_sorted = txs

    if not txs_sorted:
        c.drawString(left, y, 'No transactions')
        y -= line_height
    else:
        for date, ttype, amount, related in txs_sorted[:100]:
            if y < 80:
                new_page()
            c.drawString(left, y, str(date))
            c.drawString(left + 90, y, ttype)
            c.drawString(left + 170, y, str(amount))
            c.drawString(left + 260, y, related)
            y -= line_height

    # finish up
    c.showPage()
    c.save()
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
